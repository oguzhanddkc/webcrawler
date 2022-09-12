import pandas as pd
from openpyxl import load_workbook

file_loc = "C:/Users/oğuzhan/PycharmProjects/WebCrawler/ships.xlsx"
wb = load_workbook('ships.xlsx')
max_row = wb.active.max_row
max_col = wb.active.max_column
ship_list = []
notes = []


def read_value_from_excel(filename, column="B", row=1):
    return pd.read_excel(file_loc, skiprows=row, usecols=column, nrows=1, names=["Value"]).iloc[0]["Value"]


for x in range(max_row - 1):
    ship_list.append(read_value_from_excel("ships.xlsx", "B", x))
    notes.append(read_value_from_excel("ships.xlsx", "J", x))


def write_excel(value1, value2, value3, value4, value5, value6, value7, count):
    book = load_workbook(file_loc)
    sheet = book.active
    sheet.cell(row=count, column=3).value = value1
    sheet.cell(row=count, column=4).value = value2
    sheet.cell(row=count, column=5).value = value3
    sheet.cell(row=count, column=6).value = value4
    sheet.cell(row=count, column=7).value = value5
    sheet.cell(row=count, column=8).value = value6
    sheet.cell(row=count, column=9).value = value7

    book.save("ships.xlsx")
